
import re
import openpyxl
import math
import datetime
from dateutil.relativedelta import relativedelta


class baobiao():
    def __init__(self, date, address, fxspotare):
        self.date = datetime.datetime.strptime(date, '%Y%m%d')
        self.date0 = date
        self.address = address
        self.asset = []
        self.counterparty = {}
        self.list7fill = [0, 0, 0, 0, 0]
        self.fx = fxspotare

# 读数据
    def irsput(self):
        iaddress = self.address + '/衍生品报表/数据/irs交易查询与维护.xlsx'
        wb = openpyxl.load_workbook(iaddress)
        ws = wb.active
        for i in ws.rows:
            if i[1].value == '未到期交易':
                irstem = {}
                irstem['instrument'] = 'irs'
                irstem['code'] = i[11].value
                initialdate = datetime.datetime.strptime(
                    i[7].value.replace('-', ''), '%Y%m%d')
                irstem['initialdate'] = initialdate
                enddate = datetime.datetime.strptime(
                    i[9].value.replace('-', ''), '%Y%m%d')
                irstem['enddate'] = enddate
                irstem['initialday'] = i[10].value
                irstem['dayleft'] = (enddate - self.date).days

                if i[22].value == '上海清算所清算':
                    irstem['counterparty'] = '中央交易对手'
                else:
                    irstem['counterparty'] = i[13].value

                nextpaydate = initialdate + relativedelta(months=3)
                day = (nextpaydate - self.date).days
                while day < 0:
                    nextpaydate = nextpaydate + relativedelta(months=3)
                    day = (nextpaydate - self.date).days
                irstem['nextpaydate'] = nextpaydate

                resettype = i[40].value + i[61].value
                if resettype == '3M':
                    irstem['nextresetdate'] = nextpaydate
                elif resettype == '1W':
                    nextresetdate = initialdate + datetime.timedelta(days=7)
                    day = (nextresetdate - self.date).days

                    while day < 0:
                        nextresetdate = nextresetdate + \
                            datetime.timedelta(days=7)
                        day = (nextresetdate - self.date).days
                    irstem['nextresetdate'] = nextresetdate

                self.asset.append(irstem)
            if i[1].value == '未到期交易':
                initialdate = datetime.datetime.strptime(
                    i[3].value.replace('-', ''), '%Y%m%d')
                if initialdate.year == self.date.year:
                    self.list7fill[0] += float(i[16].value)
        iaddress = self.address + '/衍生品报表/数据/irs逐日盯市损益分析.xlsx'
        wb = openpyxl.load_workbook(iaddress)
        ws = wb.active
        for i in ws.rows:
            for j in self.asset:
                if j['code'] == i[2].value:
                    j['facevalue'] = float(i[5].value) / 10000
                    j['marketvalue'] = float(i[6].value) / 10000

                    j['moneyget'] = float(i[7].value) / 10000

                    j['moneypay'] = float(i[8].value) / 10000

                    message = re.split(r',|\(|\)|/', i[9].value)
                    j['gettype'] = message[1]
                    j['getinterest'] = message[2]
                    j['paytype'] = message[4]
                    j['payinterest'] = message[5]
                    if message[1] == 'Fix':
                        j['moneygetdate'] = j['enddate']
                        j['moneygetday'] = (j['enddate'] - self.date).days
                        j['moneypaydate'] = j['nextresetdate']
                        j['moneypayday'] = (
                            j['nextresetdate'] - self.date).days
                    else:
                        j['moneygetdate'] = j['nextresetdate']
                        j['moneygetday'] = (
                            j['nextresetdate'] - self.date).days
                        j['moneypaydate'] = j['enddate']
                        j['moneypayday'] = (j['enddate'] - self.date).days

    def crmwput(self):

        wb = openpyxl.load_workbook(
            self.address + '/衍生品报表/数据/crmw交易查询与维护.xlsx')
        ws = wb.active
        for i in ws.rows:
            if i[1].value == '未到期交易':
                crmwtem = {}
                crmwtem['instrument'] = 'crmw'
                crmwtem['code'] = i[8].value

                initialdate = datetime.datetime.strptime(
                    i[14].value.replace('-', ''), '%Y%m%d')
                crmwtem['initialdate'] = initialdate
                enddate = datetime.datetime.strptime(
                    i[19].value.replace('-', ''), '%Y%m%d')
                crmwtem['enddate'] = enddate
                crmwtem['initialday'] = (enddate - initialdate).days
                crmwtem['dayleft'] = (enddate - self.date).days
                crmwtem['counterparty'] = i[5].value
                crmwtem['type'] = i[4].value
                crmwtem['facevalue'] = float(i[16].value)

                self.asset.append(crmwtem)
            if i[1].value == '未到期交易':
                initialdate = datetime.datetime.strptime(
                    i[6].value.replace('-', ''), '%Y%m%d')
                if initialdate.year == self.date.year:
                    self.list7fill[1] += float(i[16].value) / 10000

        wb = openpyxl.load_workbook(
            self.address + '/衍生品报表/数据/crmw逐日盯市损益分析.xlsx')
        ws = wb.active
        for i in ws.rows:
            for j in self.asset:
                if j['code'] == i[2].value:
                    j['marketvalue'] = float(i[6].value) / 10000

    def forwardput(self):
        wb = openpyxl.load_workbook(
            self.address + '/衍生品报表/数据/forward逐笔损益查询.xlsx')
        ws = wb.active
        for i in ws.rows:
            if i[2].value == '远期平盘-金市-结售汇':
                enddate = datetime.datetime.strptime(
                    i[7].value.replace('-', ''), '%Y%m%d')
                if (enddate - self.date).days > 0:
                    fortem = {}
                    fortem['instrument'] = 'forward'
                    fortem['code'] = i[1].value
                    initialdate = datetime.datetime.strptime(
                        i[6].value.replace('-', ''), '%Y%m%d')
                    fortem['initialdate'] = initialdate
                    fortem['enddate'] = enddate
                    fortem['initialday'] = (enddate - initialdate).days
                    fortem['dayleft'] = (enddate - self.date).days
                    fortem['counterparty'] = i[3].value.split('(')[0]
                    fxtype = i[5].value
                    fxtype = fxtype.split('/')
                    money1 = fxtype[0]
                    money2 = fxtype[1]
                    deflator = float(i[21].value) / float(i[20].value)
                    fxrate = float(i[18].value) + float(i[19].value) / 10000
                    fxratespot = float(i[18].value)
                    value1 = float(i[9].value)
                    value2 = float(i[10].value)
                    if value1 < 0:
                        if money1 == 'USD':
                            momeypay = -value1 * deflator * fxrate
                            momeyget = value2 * deflator
                            paytype = 'USD'
                            gettype = 'CNY'
                            facevalue = value2
                            usdposition = value1
                        else:
                            momeypay = -value1 * deflator
                            momeyget = value2 * deflator * fxrate
                            paytype = 'CNY'
                            gettype = 'USD'
                            facevalue = -value1
                            usdposition = value2

                    else:
                        if money1 == 'USD':
                            momeypay = value1 * deflator * fxrate
                            momeyget = -value2 * deflator
                            paytype = 'USD'
                            gettype = 'CNY'
                            facevalue = -value2
                            usdposition = value1
                        else:
                            momeypay = value1 * deflator
                            momeyget = -value2 * deflator * fxrate
                            paytype = 'CNY'
                            gettype = 'USD'
                            facevalue = value1
                            usdposition = value1
                    fortem['facevalue'] = abs(usdposition * fxratespot / 10000)
                    fortem['marketvalue'] = float(i[21].value) / 10000
                    fortem['moneypay'] = float(momeypay) / 10000
                    fortem['paytype'] = paytype
                    fortem['moneyget'] = float(momeyget) / 10000
                    fortem['gettype'] = gettype
                    fortem['usdposition'] = usdposition * fxratespot / 10000
                    self.asset.append(fortem)
                initialdate = datetime.datetime.strptime(
                    i[6].value.replace('-', ''), '%Y%m%d')
                if initialdate.year == self.date.year:
                    self.list7fill[2] += abs(i[10].value) / 10000

    def swapput(self):
        wb = openpyxl.load_workbook(self.address + '/衍生品报表/数据/swap逐笔损益查询.xlsx')
        ws = wb.active
        for i in ws.rows:
            if i[2].value == '掉期-自营-结售汇' or i[2].value == '掉期-自营-结售汇-周游力':
                enddate = datetime.datetime.strptime(
                    i[8].value.replace('-', ''), '%Y%m%d')
                if (enddate - self.date).days > 0:
                    swaptem = {}

                    swaptem['instrument'] = 'swap'
                    swaptem['code'] = i[1].value
                    initialdate = datetime.datetime.strptime(
                        i[6].value.replace('-', ''), '%Y%m%d')
                    swaptem['initialdate'] = initialdate
                    swaptem['enddate'] = enddate
                    swaptem['initialday'] = (enddate - initialdate).days
                    swaptem['dayleft'] = (enddate - self.date).days
                    swaptem['counterparty'] = i[3].value.split('(')[0]
                    fxtype = i[5].value
                    fxtype = fxtype.split('/')
                    money1 = fxtype[0]
                    money2 = fxtype[1]
                    deflator = float(i[30].value) / float(i[27].value)
                    fxrate = float(i[31].value) + float(i[33].value) / 10000
                    fxratespot = float(i[31].value)
                    self.fx = fxratespot
                    value1 = float(i[12].value)
                    value2 = float(i[13].value)
                    if value1 < 0:
                        if money1 == 'USD':
                            momeypay = -value1 * deflator * fxrate
                            momeyget = value2 * deflator
                            paytype = 'USD'
                            gettype = 'CNY'
                            facevalue = value2
                            usdposition = value1
                        else:
                            momeypay = -value1 * deflator
                            momeyget = value2 * deflator * fxrate
                            paytype = 'CNY'
                            gettype = 'USD'
                            facevalue = -value1
                            usdposition = value2

                    else:
                        if money1 == 'USD':
                            momeypay = value1 * deflator * fxrate
                            momeyget = -value2 * deflator
                            paytype = 'USD'
                            gettype = 'CNY'
                            facevalue = -value2
                            usdposition = value1
                        else:
                            momeypay = value1 * deflator
                            momeyget = -value2 * deflator * fxrate
                            paytype = 'CNY'
                            gettype = 'USD'
                            facevalue = value1
                            usdposition = value1
                    swaptem['facevalue'] = abs(
                        usdposition * fxratespot / 10000)
                    swaptem['marketvalue'] = float(i[36].value) / 10000
                    swaptem['moneypay'] = float(momeypay) / 10000
                    swaptem['paytype'] = paytype
                    swaptem['moneyget'] = float(momeyget) / 10000
                    swaptem['gettype'] = gettype
                    swaptem['usdposition'] = usdposition * fxratespot / 10000

                    self.asset.append(swaptem)
                initialdate = datetime.datetime.strptime(
                    i[6].value.replace('-', ''), '%Y%m%d')
                if initialdate.year == self.date.year:
                    self.list7fill[3] += abs(i[12].value) * self.fx / 10000

    def optionput(self):
        wb = openpyxl.load_workbook(
            self.address + '/衍生品报表/数据/option逐笔损益查询.xlsx')
        ws = wb.active
        for i in ws.rows:
            if i[4].value == '期权-自营-结售汇':
                enddate = datetime.datetime.strptime(
                    i[13].value.replace('-', ''), '%Y%m%d')
                dayleft = (enddate - self.date).days
                if dayleft > 0:
                    opttem = {}
                    opttem['instrument'] = 'option'
                    opttem['code'] = i[1].value
                    opttem['initialdate'] = datetime.datetime.strptime(
                        i[8].value.replace('-', ''), '%Y%m%d')
                    opttem['enddate'] = enddate
                    opttem['initialday'] = (
                        enddate - opttem['initialdate']).days
                    opttem['dayleft'] = dayleft
                    opttem['counterparty'] = i[5].value.split('(')[0]
                    opttem['type'] = i[10].value
                    fxratespot = self.fx
                    opttem['facevalue'] = float(
                        i[18].value) / 10000 * fxratespot
                    opttem['marketvalue'] = float(i[45].value) / 10000

                    opttem['fundamental'] = float(
                        i[18].value) * fxratespot / 10000
                    if i[47].value == '':
                        delta = 0
                    else:
                        delta = abs(float(i[47].value) / float(i[19].value))
                    opttem['usd'] = float(
                        i[18].value) * fxratespot * delta / 10000
                    opttem['cny'] = float(i[19].value) * delta / 10000
                    self.asset.append(opttem)
                initialdate = datetime.datetime.strptime(
                    i[8].value.replace('-', ''), '%Y%m%d')
                if initialdate.year == self.date.year:
                    self.list7fill[4] += abs(float(i[18].value) /
                                             10000 * self.fx)

# 写报表
    def fillcontent(self, filler, xlsindex, xlscontent):
        if xlsindex not in filler.keys():  # 如果这个坐标没有值，就填进去
            filler[xlsindex] = xlscontent
        else:
            filler[xlsindex] += xlscontent  # 如果这个坐标有值，就将原来的值和新值相加填进去

    def g4c_1bfiller(self, filler, position, value, day, interest):
        upper1 = [
            30,
            91,
            183,
            365,
            730,
            1095,
            1460,
            1825,
            2555,
            3650,
            5475,
            7300,
            36000,
            36000]
        upper2 = [
            30,
            91,
            183,
            365,
            694,
            1022,
            1314,
            1570,
            2081,
            2665,
            3395,
            3869,
            4380,
            7300]
        if position == 'long':
            columnindex = '8'
        else:
            columnindex = '9'
        if float(interest) < 3:
            upper = upper2
        else:
            upper = upper1
        if day <= upper[0]:
            self.fillcontent(filler, '7,' + columnindex, value)
        elif day <= upper[1]:
            self.fillcontent(filler, '8,' + columnindex, value)
        elif day <= upper[2]:
            self.fillcontent(filler, '9,' + columnindex, value)
        elif day <= upper[3]:
            self.fillcontent(filler, '10,' + columnindex, value)
        elif day <= upper[4]:
            self.fillcontent(filler, '11,' + columnindex, value)
        elif day <= upper[5]:
            self.fillcontent(filler, '12,' + columnindex, value)
        elif day <= upper[6]:
            self.fillcontent(filler, '13,' + columnindex, value)
        elif day <= upper[7]:
            self.fillcontent(filler, '14,' + columnindex, value)
        elif day <= upper[8]:
            self.fillcontent(filler, '15,' + columnindex, value)
        elif day <= upper[9]:
            self.fillcontent(filler, '16,' + columnindex, value)
        elif day <= upper[10]:
            self.fillcontent(filler, '17,' + columnindex, value)
        elif day <= upper[11]:
            self.fillcontent(filler, '18,' + columnindex, value)
        elif day <= upper[12]:
            self.fillcontent(filler, '19,' + columnindex, value)
        elif day <= upper[13]:
            self.fillcontent(filler, '20,' + columnindex, value)
        elif day <= upper[14]:
            self.fillcontent(filler, '21,' + columnindex, value)

    def g4c_1b(self):
        address1 = self.address + '/衍生品报表/报表/G4C-1(b)人民币.xlsx'
        address2 = self.address + '/衍生品报表/报表/G4C-1(b)美元.xlsx'
        address11 = self.address + '/衍生品报表/结果/G4C-1(b)人民币-' + date + '.xlsx'
        address22 = self.address + '/衍生品报表/结果/G4C-1(b)美元-' + date + '.xlsx'
        filler1 = {}
        filler2 = {}
        for i in self.asset:
            if i['instrument'] == 'irs':
                self.g4c_1bfiller(
                    filler1,
                    'long',
                    i['moneyget'],
                    i['moneygetday'],
                    i['getinterest'])
                self.g4c_1bfiller(
                    filler1,
                    'short',
                    i['moneypay'],
                    i['moneypayday'],
                    i['payinterest'])
            if i['instrument'] == 'forward' or i['instrument'] == 'swap':
                if i['paytype'] == 'USD':
                    self.g4c_1bfiller(
                        filler1, 'long', i['moneyget'], i['dayleft'], 3)
                    self.g4c_1bfiller(
                        filler2, 'short', i['moneypay'], i['dayleft'], 3)
                else:
                    self.g4c_1bfiller(
                        filler2, 'long', i['moneyget'], i['dayleft'], 3)
                    self.g4c_1bfiller(
                        filler1, 'short', i['moneypay'], i['dayleft'], 3)
            """
            if i['instrument']=='option':
                if i['type']=='PUT':
                    self.g4c_1bfiller(filler1, 'long', i['cny'], i['dayleft'], 3)
                    self.g4c_1bfiller(filler2, 'short', i['usd'], i['dayleft'], 3)
                else:
                    self.g4c_1bfiller(filler1, 'short', i['cny'], i['dayleft'], 3)
                    self.g4c_1bfiller(filler2, 'long', i['usd'], i['dayleft'], 3)
            """

        wb = openpyxl.load_workbook(address1)
        ws = wb.active
        for a, b in filler1.items():
            xy = a.split(',')
            ws.cell(row=int(xy[0]), column=int(xy[1])).value = b

        wb.save(address11)

        wb = openpyxl.load_workbook(address2)
        ws = wb.active
        for a, b in filler2.items():
            xy = a.split(',')
            ws.cell(row=int(xy[0]), column=int(xy[1])).value = b
        wb.save(address22)

    def g4b_3(self):  # 顺便把G44也填了
        address1 = self.address + '/衍生品报表/报表/G4B-3(a).xlsx'
        address2 = self.address + '/衍生品报表/报表/G4B-3.xlsx'
        address3 = self.address + '/衍生品报表/报表/G44.xlsx'
        address11 = self.address + '/衍生品报表/结果//G4B-3(a)-' + date + '.xlsx'
        address22 = self.address + '/衍生品报表/结果/G4B-3-' + date + '.xlsx'
        address33 = self.address + '/衍生品报表/结果/G44-' + date + '.xlsx'
        filler1 = {}
        filler2 = {}
        counterpaty = {}
        wilist = {
            'AAA': 0.007,
            'AA': 0.007,
            'A': 0.008,
            'BBB': 0.01,
            'BB': 0.02,
            'B': 0.03,
            'CCC': 0.1}
        crisklist = {
            '0': 0,
            '0.2': 1,
            '0.25': 2,
            '0.5': 3,
            '0.75': 4,
            '1': 5,
            '1.5': 6}
        for i in self.asset:
            if i['counterparty'] not in counterpaty.keys():
                if i['counterparty'] == '中央交易对手':
                    counterpaty[i['counterparty']] = [
                        'NOTBANK', 'CLAERING', 'AAA', 0.02, 0, 0, 0, 0]
                elif '银行' in i['counterparty']:
                    counterpaty[i['counterparty']] = [
                        'BANK', 'NOTCLEARING', 'AAA', 0.2, 0, 0, 0, 0]
        A = 0
        B = 0
        g44one = 0
        g44two = 0
        for ctp, message in counterpaty.items():
            clearingno = 0
            if message[1] == 'NOTCLEARING':
                for asset in self.asset:
                    if asset['counterparty'] == ctp:
                        if asset['instrument'] == 'irs':
                            row = 29
                            rowplus = 0
                            message[6] = max(
                                message[6], float(
                                    asset['dayleft']) / 365)

                            if asset['dayleft'] <= 365:
                                addratio = 0

                            elif asset['dayleft'] <= 1825:
                                addratio = 0.005
                                rowplus += 8
                            else:
                                addratio = 0.01
                                rowplus += 16
                            if message[0] == 'NOTBANK':
                                riskratio = message[3]
                            else:
                                if asset['initialday'] <= 95:
                                    riskratio = 0.2
                                else:
                                    riskratio = 0.25
                            rowplus += crisklist[str(riskratio)]

                            ead = asset['facevalue'] * addratio + \
                                max(asset['marketvalue'], 0)
                            riskasset = ead * riskratio
                            message[4] += ead
                            message[5] += riskasset
                        elif asset['instrument'] == 'forward' or asset['instrument'] == 'option' or asset['instrument'] == 'swap':
                            row = 54
                            rowplus = 0
                            message[6] = max(
                                message[6], float(
                                    asset['dayleft']) / 365)

                            if asset['dayleft'] <= 365:
                                addratio = 0.01
                            elif asset['dayleft'] <= 1825:
                                addratio = 0.05
                                rowplus += 8
                            else:
                                addratio = 0.075
                                rowplus += 16
                            if message[0] == 'NOTBANK':
                                riskratio = message[3]
                            else:
                                if asset['initialday'] <= 92:
                                    riskratio = 0.2
                                else:
                                    riskratio = 0.25
                            rowplus += crisklist[str(riskratio)]

                            ead = asset['facevalue'] * addratio + \
                                max(asset['marketvalue'], 0)
                            riskasset = ead * riskratio
                            message[4] += ead
                            message[5] += riskasset
                        elif asset['instrument'] == 'crmw':

                            row = 12
                            rowplus = 0
                            message[6] = max(
                                message[6], float(
                                    asset['dayleft']) / 365)

                            addratio = 0.05

                            if message[0] == 'NOTBANK':
                                riskratio = message[3]
                            else:
                                if asset['initialday'] <= 95:
                                    riskratio = 0.2
                                else:
                                    riskratio = 0.25
                            rowplus += crisklist[str(riskratio)]

                            ead = asset['facevalue'] * addratio + \
                                max(asset['marketvalue'], 0)
                            riskasset = ead * riskratio
                            message[4] += ead
                            message[5] += riskasset
                            self.fillcontent(filler1, str(
                                row + rowplus) + ',6', ead)
                        else:
                            continue

                        self.fillcontent(
                            filler1, str(
                                row + rowplus) + ',3', asset['facevalue'])
                        self.fillcontent(filler1, str(
                            row + rowplus) + ',5', max(asset['marketvalue'], 0))

                        self.fillcontent(filler2, '6,3', asset['facevalue'])
                        self.fillcontent(filler2, '6,4', riskasset)
                        g44one += asset['facevalue'] * addratio
                        g44two += max(asset['marketvalue'], 0)

            else:
                ngrup = 0
                ngrdown = 0
                agross = 0
                dayplusfacevalue = 0
                facevalueplus = 0

                for asset in self.asset:
                    if asset['counterparty'] == ctp:
                        if asset['instrument'] == 'irs':
                            dayplusfacevalue += asset['facevalue'] * \
                                asset['dayleft']
                            facevalueplus += asset['facevalue']
                            if asset['dayleft'] <= 365:
                                addratio = 0
                            elif asset['dayleft'] <= 1825:
                                addratio = 0.005
                            else:
                                addratio = 0.01
                            if message[0] == 'BANK' and asset['dayleft'] > 91:
                                message[3] = 0.25
                            ngrup += asset['marketvalue']
                            ngrdown += max(asset['marketvalue'], 0)
                            agross += asset['facevalue'] * addratio

                        elif asset['instrument'] == 'forward' or asset['instrument'] == 'option' or asset['instrument'] == 'swap':
                            dayplusfacevalue += asset['facevalue'] * \
                                asset['dayleft']
                            facevalueplus += asset['facevalue']
                            if asset['dayleft'] <= 365:
                                addratio = 0.01
                            elif asset['dayleft'] <= 1825:
                                addratio = 0.05
                            else:
                                addratio = 0.075
                            if message[0] == 'BANK' and asset['dayleft'] > 91:
                                message[3] = 0.25
                            ngrup += asset['marketvalue']
                            ngrdown += max(asset['marketvalue'], 0)
                            agross += asset['facevalue'] * addratio
                        elif asset['instrument'] == 'crmw':
                            dayplusfacevalue += asset['facevalue'] * \
                                asset['dayleft']
                            facevalueplus += asset['facevalue']
                            addratio = 0.05
                            if message[0] == 'BANK' and asset['dayleft'] > 91:
                                message[3] = 0.25
                            ngrup += asset['marketvalue']
                            ngrdown += max(asset['marketvalue'], 0)
                            agross += asset['facevalue'] * addratio

                if ngrdown == 0:
                    ngr = 0
                else:
                    ngr = max(ngrup, 0) / ngrdown
                ead = 0.4 * agross + 0.6 * agross * ngr + max(ngrup, 0)
                riskasset = ead * message[3]
                message[4] = ead
                message[5] = riskasset

                message[6] = dayplusfacevalue / facevalueplus / 365
                g44one += 0.4 * agross + 0.6 * agross * ngr
                g44two += max(ngrup, 0)
                if ctp == '中央交易对手':
                    self.fillcontent(filler2, '9,3', facevalueplus)
                    self.fillcontent(filler2, '9,4', riskasset)
                else:

                    self.fillcontent(
                        filler1, str(
                            169 + clearingno) + ',2', ctp)
                    self.fillcontent(
                        filler1, str(
                            169 + clearingno) + ',3', facevalueplus)
                    self.fillcontent(
                        filler1, str(
                            169 + clearingno) + ',4', agross)
                    self.fillcontent(
                        filler1, str(
                            169 + clearingno) + ',5', ngr)
                    self.fillcontent(filler1, str(
                        169 + clearingno) + ',7', max(ngrup, 0))
                    self.fillcontent(
                        filler1, str(
                            169 + clearingno) + ',9', message[3])
                    self.fillcontent(filler2, '6,3', facevalueplus)
                    self.fillcontent(filler2, '6,4', riskasset)
                    clearingno += 1

            wi = wilist[message[2]]
            deflator = (1 - math.exp(-0.05 * message[6])) / (0.05 * message[6])
            message[7] = message[4] * deflator * message[6] * wi
            if ctp != '中央交易对手':
                A += message[7]
                B += message[7] * message[7]

        CVA = 12.5 * 2.33 * math.sqrt(0.25 * A * A + 0.75 * B)

        self.fillcontent(filler1, '180,3', CVA)
        self.fillcontent(filler2, '7,4', CVA)
        wb = openpyxl.load_workbook(address1)
        ws = wb.active
        for a, b in filler1.items():
            xy = a.split(',')
            ws.cell(row=int(xy[0]), column=int(xy[1])).value = b
        wb.save(address11)

        wb = openpyxl.load_workbook(address2)
        ws = wb.active
        for a, b in filler2.items():
            xy = a.split(',')
            ws.cell(row=int(xy[0]), column=int(xy[1])).value = b
        wb.save(address22)

        wb = openpyxl.load_workbook(address3)
        ws = wb.active

        ws.cell(row=12, column=3).value = g44two
        ws.cell(row=13, column=3).value = g44one
        wb.save(address33)

        print('------------------g4b-3相关信息-----------------')
        print('>>数据含义：[是否银行，是否中央交易，外部评级，风险权重，风险暴露，风险资产，剩余期限，CVA计算量]')
        for a, b in counterpaty.items():
            print(a, b)

    def g4c_1a(self):
        address = self.address + '/衍生品报表/报表/G4C-1(a).xlsx'
        address0 = self.address + '/衍生品报表/结果/G4C-1(a)-' + date + '.xlsx'
        filler = {}

        for asset in self.asset:
            if asset['instrument'] == 'crmw':
                if asset['type'] == '买入':
                    row = 21
                else:
                    row = 20
                if asset['dayleft'] <= 183:
                    self.fillcontent(
                        filler, str(row) + ',4', asset['facevalue'])
                elif asset['dayleft'] <= 730:
                    self.fillcontent(
                        filler, str(row) + ',5', asset['facevalue'])
                else:
                    self.fillcontent(
                        filler, str(row) + ',6', asset['facevalue'])
        wb = openpyxl.load_workbook(address)
        ws = wb.active
        for a, b in filler.items():
            xy = a.split(',')
            ws.cell(row=int(xy[0]), column=int(xy[1])).value = b
        wb.save(address0)

    def g4c_1e(self):
        address = self.address + '/衍生品报表/报表/G4C-1(e).xlsx'
        address0 = self.address + '/衍生品报表/结果/G4C-1(e)-' + date + '.xlsx'
        filler = {}
        for asset in self.asset:
            if asset['instrument'] == 'forward' or asset['instrument'] == 'swap':
                self.fillcontent(filler, '6,3', asset['usdposition'])
        wb = openpyxl.load_workbook(address)
        ws = wb.active
        for a, b in filler.items():
            xy = a.split(',')
            ws.cell(row=int(xy[0]), column=int(xy[1])).value = b
        wb.save(address0)

    def g4c_1h(self):
        address = self.address + '/衍生品报表/报表/G4C-1(h).xlsx'
        address0 = self.address + '/衍生品报表/结果/G4C-1(h)-' + date + '.xlsx'
        filler = {}
        for asset in self.asset:
            if asset['instrument'] == 'option':
                if asset['type'] == 'CALL':
                    column = 6
                elif asset['type'] == 'PUT':
                    column = 5
                self.fillcontent(
                    filler,
                    '14,' + str(column),
                    asset['marketvalue'])
                riskasset = max(
                    min(asset['marketvalue'], asset['fundamental'] * 0.08), 0)
                self.fillcontent(filler, '14,8', riskasset)
        wb = openpyxl.load_workbook(address)
        ws = wb.active
        for a, b in filler.items():
            xy = a.split(',')
            ws.cell(row=int(xy[0]), column=int(xy[1])).value = b
        wb.save(address0)

    def a1411(self):
        address = self.address + '/衍生品报表/报表/A1411金融机构资产负债项目月报表.xlsx'
        address0 = self.address + '/衍生品报表/结果/A1411金融机构资产负债项目月报表-' + date + '.xlsx'
        filler = {}
        irs = 0
        crmw = 0

        for i in self.asset:
            if i['instrument'] == 'irs':
                irs += i['marketvalue']
            elif i['instrument'] == 'crmw':
                crmw += i['marketvalue']
            elif i['instrument'] == 'forward':
                if i['marketvalue'] < 0:
                    self.fillcontent(filler, '12,4', -i['marketvalue'])
                else:
                    self.fillcontent(filler, '6,4', i['marketvalue'])
            elif i['instrument'] == 'option':
                self.fillcontent(filler, '8,4', i['marketvalue'])
            elif i['instrument'] == 'swap':
                if i['marketvalue'] < 0:
                    self.fillcontent(filler, '15,4', -i['marketvalue'])
                else:
                    self.fillcontent(filler, '9,4', i['marketvalue'])
        if irs < 0:
            self.fillcontent(filler, '15,4', -irs)
        else:
            self.fillcontent(filler, '9,4')
        self.fillcontent(filler, '9,4', crmw)
        wb = openpyxl.load_workbook(address)
        ws = wb.active
        for a, b in filler.items():
            xy = a.split(',')
            ws.cell(row=int(xy[0]), column=int(xy[1])).value = b
        wb.save(address0)

    def g21(self):
        address = self.address + '/衍生品报表/报表/G21.xlsx'
        address0 = self.address + '/衍生品报表/结果/G21-' + date + '.xlsx'
        filler = {}
        irscolumns = []
        irs = 0

        for i in self.asset:
            if i['dayleft'] <= 1:
                column = 3
            elif i['dayleft'] <= 7:
                column = 4
            elif i['dayleft'] <= 30:
                column = 5
            elif i['dayleft'] <= 90:
                column = 6
            elif i['dayleft'] <= 365:
                column = 7
            elif i['dayleft'] <= 1826:
                column = 8
            elif i['dayleft'] <= 3652:
                column = 9

            if i['marketvalue'] < 0:
                row = 38
            else:
                row = 21
            if i['instrument'] == 'irs':
                irscolumns.append([column, i['marketvalue']])
                irs += i['marketvalue']
            else:
                self.fillcontent(filler, str(row) + ',' +
                                 str(column), abs(i['marketvalue']))
        if irs < 0:
            irsrow = 38
            reverseindex = -1
        else:
            irsrow = 21
            reverseindex = 1

        for i in irscolumns:
            self.fillcontent(filler, str(irsrow) + ',' +
                             str(i[0]), i[1] * reverseindex)

        wb = openpyxl.load_workbook(address)
        ws = wb.active
        for a, b in filler.items():
            xy = a.split(',')
            ws.cell(row=int(xy[0]), column=int(xy[1])).value = b
        wb.save(address0)

    def g01(self):
        address = self.address + '/衍生品报表/报表/G01_I.xlsx'
        address0 = self.address + '/衍生品报表/结果/G01_I-' + date + '.xlsx'

        facevalue = 0
        facevalue0 = [0, 0, 0, 0, 0]

        for i in self.asset:
            facevalue += i['facevalue']
            if i['instrument'] == 'irs':
                facevalue0[0] += i['facevalue']
            elif i['instrument'] == 'crmw':
                facevalue0[1] += i['facevalue']
            elif i['instrument'] == 'forward':
                facevalue0[2] += i['facevalue']
            elif i['instrument'] == 'option':
                facevalue0[3] += i['facevalue']
            elif i['instrument'] == 'swap':
                facevalue0[4] += i['facevalue']
        print('---------------g01相关信息----------------')
        print('总数：' + str(facevalue))
        print('利率互换：' + str(facevalue0[0]))
        print('CRMW：' + str(facevalue0[1]))
        print('外汇远期：' + str(facevalue0[2]))
        print('外汇期权：' + str(facevalue0[3]))
        print('外汇掉期：' + str(facevalue0[4]))
        wb = openpyxl.load_workbook(address)
        ws = wb.active
        ws.cell(column=4, row=42).value = facevalue
        wb.save(address0)

    def g02(self):
        address = self.address + '/衍生品报表/报表/G02.xlsx'
        address0 = self.address + '/衍生品报表/结果/G02-' + date + '.xlsx'
        filler = {}
        for i in self.asset:
            if i['instrument'] == 'irs':
                self.fillcontent(filler, '11,4', i['facevalue'])
                self.fillcontent(filler, '11,5', i['facevalue'])
                self.fillcontent(filler, '46,4', i['facevalue'])
                self.fillcontent(filler, '46,5', i['marketvalue'])
                if i['marketvalue'] > 0:
                    self.fillcontent(filler, '30,4', i['marketvalue'])
                else:
                    self.fillcontent(filler, '30,5', abs(i['marketvalue']))
            elif i['instrument'] == 'crmw':
                self.fillcontent(filler, '11,12', i['facevalue'])
                self.fillcontent(filler, '11,13', i['facevalue'])
                self.fillcontent(filler, '46,12', i['facevalue'])
                self.fillcontent(filler, '46,13', i['marketvalue'])
                if i['marketvalue'] > 0:
                    self.fillcontent(filler, '30,12', i['marketvalue'])
                else:
                    self.fillcontent(filler, '30,13', abs(i['marketvalue']))
            elif i['instrument'] == 'option':
                self.fillcontent(filler, '7,6', i['facevalue'])
                self.fillcontent(filler, '7,7', i['facevalue'])
                self.fillcontent(filler, '46,6', i['facevalue'])
                self.fillcontent(filler, '46,7', i['marketvalue'])
                if i['marketvalue'] > 0:
                    self.fillcontent(filler, '26,6', i['marketvalue'])
                else:
                    self.fillcontent(filler, '26,7', abs(i['marketvalue']))
            elif i['instrument'] == 'forward':
                self.fillcontent(filler, '10,6', i['facevalue'])
                self.fillcontent(filler, '10,7', i['facevalue'])
                self.fillcontent(filler, '46,6', i['facevalue'])
                self.fillcontent(filler, '46,7', i['marketvalue'])
                if i['marketvalue'] > 0:
                    self.fillcontent(filler, '29,6', i['marketvalue'])
                else:
                    self.fillcontent(filler, '29,7', abs(i['marketvalue']))
            elif i['instrument'] == 'swap':
                self.fillcontent(filler, '11,6', i['facevalue'])
                self.fillcontent(filler, '11,7', i['facevalue'])
                self.fillcontent(filler, '46,6', i['facevalue'])
                self.fillcontent(filler, '46,7', i['marketvalue'])
                if i['marketvalue'] > 0:
                    self.fillcontent(filler, '30,6', i['marketvalue'])
                else:
                    self.fillcontent(filler, '30,7', abs(i['marketvalue']))
        positive = 0
        negative = 0
        for m, n in filler.items():
            if m.split(',')[0] == '46' and int(m.split(',')[1]) % 2 == 1:
                positive += max(0, n)
                negative += min(0, n)
        self.fillcontent(filler, '46,19', positive)
        self.fillcontent(filler, '46,20', negative)
        wb = openpyxl.load_workbook(address)
        ws = wb.active
        for a, b in filler.items():
            xy = a.split(',')
            ws.cell(row=int(xy[0]), column=int(xy[1])).value = b
        wb.save(address0)

    def g33filler(self, filler, position, value, day, type):

        if type == 'forward':
            row = 25
        elif type == 'irs':
            row = 27
        elif type == 'swap':
            row = 29
        elif type == 'option':
            row = 33
        if position == 'long':
            rowplus = 0
        else:
            rowplus = 1
        row += rowplus

        if day <= 30:
            column = 4
        elif day <= 91:
            column = 5
        elif day <= 183:
            column = 6
        elif day <= 365:
            column = 7
        elif day <= 730:
            column = 8
        elif day <= 1095:
            column = 9
        elif day <= 1460:
            column = 10
        elif day <= 1825:
            column = 11
        elif day <= 2555:
            column = 12
        elif day <= 3650:
            column = 13
        elif day <= 7300:
            column = 15
        self.fillcontent(filler, str(row) + ',' + str(column), value)

    def g33(self):
        address1 = self.address + '/衍生品报表/报表/G33_I人民币.xlsx'
        address2 = self.address + '/衍生品报表/报表/G33_I美元.xlsx'
        address11 = self.address + '/衍生品报表/结果/G33_I人民币-' + date + '.xlsx'
        address22 = self.address + '/衍生品报表/结果/G33_I美元-' + date + '.xlsx'

        filler1 = {}
        filler2 = {}
        for i in self.asset:
            if i['instrument'] == 'irs':

                self.g33filler(
                    filler1,
                    'long',
                    i['moneyget'],
                    i['moneygetday'],
                    i['instrument'])
                self.g33filler(
                    filler1,
                    'short',
                    i['moneypay'],
                    i['moneypayday'],
                    i['instrument'])
            if i['instrument'] == 'forward' or i['instrument'] == 'swap':
                if i['paytype'] == 'USD':
                    self.g33filler(
                        filler1,
                        'long',
                        i['moneyget'],
                        i['dayleft'],
                        i['instrument'])
                    self.g33filler(
                        filler2,
                        'short',
                        i['moneypay'],
                        i['dayleft'],
                        i['instrument'])
                else:
                    self.g33filler(
                        filler2,
                        'long',
                        i['moneyget'],
                        i['dayleft'],
                        i['instrument'])
                    self.g33filler(
                        filler1,
                        'short',
                        i['moneypay'],
                        i['dayleft'],
                        i['instrument'])
            if i['instrument'] == 'option':
                if i['type'] == 'PUT':
                    self.g33filler(
                        filler1,
                        'long',
                        i['cny'],
                        i['dayleft'],
                        i['instrument'])
                    self.g33filler(
                        filler2,
                        'short',
                        i['usd'],
                        i['dayleft'],
                        i['instrument'])
                else:
                    self.g33filler(
                        filler1,
                        'short',
                        i['cny'],
                        i['dayleft'],
                        i['instrument'])
                    self.g33filler(
                        filler2,
                        'long',
                        i['usd'],
                        i['dayleft'],
                        i['instrument'])

        wb = openpyxl.load_workbook(address1)
        ws = wb.active
        for a, b in filler1.items():
            xy = a.split(',')
            ws.cell(row=int(xy[0]), column=int(xy[1])).value = b

        wb.save(address11)

        wb = openpyxl.load_workbook(address2)
        ws = wb.active
        for a, b in filler2.items():
            xy = a.split(',')
            ws.cell(row=int(xy[0]), column=int(xy[1])).value = b
        wb.save(address22)

    def list7(self):
        address = self.address + '/衍生品报表/报表/表7衍生产品交易业务统计表.xlsx'
        address0 = self.address + '/衍生品报表/结果/表7衍生产品交易业务统计表-' + date + '.xlsx'
        filler = {}
        for i in self.asset:
            if i['instrument'] == 'irs':
                self.fillcontent(filler, '11,4', i['facevalue'])
                self.fillcontent(filler, '11,5', i['facevalue'])
                self.fillcontent(filler, '65,4', i['facevalue'])
                self.fillcontent(filler, '65,5', i['marketvalue'])
                if i['marketvalue'] > 0:
                    self.fillcontent(filler, '30,4', i['marketvalue'])
                else:
                    self.fillcontent(filler, '30,5', abs(i['marketvalue']))
            elif i['instrument'] == 'crmw':
                self.fillcontent(filler, '11,12', i['facevalue'])
                self.fillcontent(filler, '11,13', i['facevalue'])
                self.fillcontent(filler, '65,12', i['facevalue'])
                self.fillcontent(filler, '65,13', i['marketvalue'])
                if i['marketvalue'] > 0:
                    self.fillcontent(filler, '30,12', i['marketvalue'])
                else:
                    self.fillcontent(filler, '30,13', abs(i['marketvalue']))
            elif i['instrument'] == 'option':
                self.fillcontent(filler, '7,6', i['facevalue'])
                self.fillcontent(filler, '7,7', i['facevalue'])
                self.fillcontent(filler, '65,6', i['facevalue'])
                self.fillcontent(filler, '65,7', i['marketvalue'])
                if i['marketvalue'] > 0:
                    self.fillcontent(filler, '26,6', i['marketvalue'])
                else:
                    self.fillcontent(filler, '26,7', abs(i['marketvalue']))
            elif i['instrument'] == 'forward':
                self.fillcontent(filler, '10,6', i['facevalue'])
                self.fillcontent(filler, '10,7', i['facevalue'])
                self.fillcontent(filler, '65,6', i['facevalue'])
                self.fillcontent(filler, '65,7', i['marketvalue'])
                if i['marketvalue'] > 0:
                    self.fillcontent(filler, '29,6', i['marketvalue'])
                else:
                    self.fillcontent(filler, '29,7', abs(i['marketvalue']))
            elif i['instrument'] == 'swap':
                self.fillcontent(filler, '11,6', i['facevalue'])
                self.fillcontent(filler, '11,7', i['facevalue'])
                self.fillcontent(filler, '65,6', i['facevalue'])
                self.fillcontent(filler, '65,7', i['marketvalue'])
                if i['marketvalue'] > 0:
                    self.fillcontent(filler, '30,6', i['marketvalue'])
                else:
                    self.fillcontent(filler, '30,7', abs(i['marketvalue']))
        positive = 0
        negative = 0
        for m, n in filler.items():
            if m.split(',')[0] == '65' and int(m.split(',')[1]) % 2 == 1:
                positive += max(0, n)
                negative += min(0, n)
        self.fillcontent(filler, '65,19', positive)
        self.fillcontent(filler, '65,20', negative)

        self.fillcontent(filler, '49,4', self.list7fill[0])
        self.fillcontent(filler, '49,5', self.list7fill[0])
        self.fillcontent(filler, '49,12', self.list7fill[1])
        self.fillcontent(filler, '49,13', self.list7fill[1])
        self.fillcontent(filler, '48,6', self.list7fill[2])
        self.fillcontent(filler, '48,7', self.list7fill[2])
        self.fillcontent(filler, '49,6', self.list7fill[3])
        self.fillcontent(filler, '49,7', self.list7fill[3])
        self.fillcontent(filler, '45,6', self.list7fill[4])
        self.fillcontent(filler, '45,7', self.list7fill[4])
        wb = openpyxl.load_workbook(address)
        ws = wb.active
        for a, b in filler.items():
            xy = a.split(',')
            ws.cell(row=int(xy[0]), column=int(xy[1])).value = b
        wb.save(address0)

    def g22(self):  # g501也一起填
        address1 = self.address + '/衍生品报表/报表/G22.xlsx'
        address11 = self.address + '/衍生品报表/结果/G22-' + date + '.xlsx'
        address2 = self.address + '/衍生品报表/报表/G25_I.xlsx'
        address22 = self.address + '/衍生品报表/结果/G25_I-' + date + '.xlsx'

        domasset = 0
        domdebt = 0
        forasset = 0
        fordebt = 0
        for i in self.asset:
            if i['dayleft'] <= 31:
                if i['instrument'] in ['irs', 'crmw']:
                    domasset += max(i['marketvalue'], 0)
                    domdebt += min(i['marketvalue'], 0)
                else:
                    forasset += max(i['marketvalue'], 0)
                    fordebt += min(i['marketvalue'], 0)
        moneyin = domasset + forasset
        moneyout = abs(domdebt + fordebt)
        wb = openpyxl.load_workbook(address1)
        ws = wb.active

        ws.cell(row=14, column=3).value = moneyin

        ws.cell(row=23, column=3).value = moneyout

        wb.save(address11)

        wb = openpyxl.load_workbook(address2)
        ws = wb.active

        ws.cell(row=83, column=3).value = moneyout
        ws.cell(row=149, column=3).value = moneyin

        wb.save(address22)


date = '20200331'
address = 'C:/Users/zyzse/Desktop'
fxspotare = 7.0851
test = baobiao(date, address, fxspotare)
test.irsput()
test.crmwput()
test.forwardput()
test.optionput()
test.swapput()


test.g4c_1b()
test.g4b_3()
test.g4c_1a()
test.g4c_1e()
test.g4c_1h()
test.a1411()
test.g21()
test.g01()
test.g02()
test.g33()
test.list7()
test.g22()
print('---------------资产信息----------------')
for a in test.asset:
    print(a)
